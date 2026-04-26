from pathlib import Path

from openpyxl import load_workbook

from app.core.providers.base import ExcelProvider, SheetData


class LocalFileProvider(ExcelProvider):
    """Reads/writes Excel files stored on the local filesystem.

    Supports .xlsx, .xlsm files. For .xls (legacy binary format), open in Excel
    and save as .xlsx first. Works for files anywhere on disk including
    OneDrive/Dropbox/iCloud sync folders (treated as plain local paths).

    The `access_token` parameter is ignored — local files don't need auth.
    """

    @staticmethod
    def _resolve_path(file_url: str) -> Path:
        # Strip optional file:// scheme
        path_str = file_url.removeprefix("file://")
        path = Path(path_str).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {path}")
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ValueError(
                f"Unsupported file format '{path.suffix}'. "
                f"Open the file in Excel and save as .xlsx, then update EXCEL_FILE_URL."
            )
        return path

    async def read_sheet(self, file_url: str, sheet_name: str, access_token: str) -> SheetData:
        path = self._resolve_path(file_url)
        wb = load_workbook(path, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            ws = wb[sheet_name]

            all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
            if not all_rows:
                return SheetData(headers=[], rows=[])

            headers = [str(h) if h is not None else "" for h in all_rows[0]]
            rows = all_rows[1:]
            return SheetData(headers=headers, rows=rows)
        finally:
            wb.close()

    async def update_cell(
        self, file_url: str, sheet_name: str, cell_address: str, value: str, access_token: str
    ) -> None:
        path = self._resolve_path(file_url)
        wb = load_workbook(path)
        try:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            ws = wb[sheet_name]
            ws[cell_address] = value
            wb.save(path)
        finally:
            wb.close()
